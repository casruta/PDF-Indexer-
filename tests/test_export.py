"""Tests for analyst-friendly export formats."""

from __future__ import annotations

import csv
import io
import json
import zipfile

import pytest

from pdf_indexer.export import (
    generate_table_csv,
    generate_combined_csv,
    generate_json_export,
    generate_excel,
    generate_zip_bundle,
    generate_data_quality_report,
)


@pytest.fixture
def sample_extraction():
    """Multi-table extraction result with mixed data types."""
    return {
        "metadata": {"title": "Test Report", "author": "Analyst", "page_count": "2"},
        "pages": [
            {
                "page_number": 1,
                "tables": [
                    {
                        "table_number": 1,
                        "headers": ["Segment", "Q3 2025", "Growth"],
                        "rows": [
                            [
                                {"value": "North America", "data_type": "text", "numeric_value": None},
                                {"value": "$12,450,000", "data_type": "currency", "numeric_value": 12450000.0},
                                {"value": "15.3%", "data_type": "percent", "numeric_value": 15.3},
                            ],
                            [
                                {"value": "Europe", "data_type": "text", "numeric_value": None},
                                {"value": "$8,750,000", "data_type": "currency", "numeric_value": 8750000.0},
                                {"value": "(2.1%)", "data_type": "percent", "numeric_value": -2.1},
                            ],
                        ],
                        "row_count": 2,
                        "col_count": 3,
                        "table_type": "financial",
                    }
                ],
            },
            {
                "page_number": 2,
                "tables": [
                    {
                        "table_number": 2,
                        "headers": ["Region", "Units Sold"],
                        "rows": [
                            [
                                {"value": "West", "data_type": "text", "numeric_value": None},
                                {"value": "1,250", "data_type": "number", "numeric_value": 1250.0},
                            ],
                        ],
                        "row_count": 1,
                        "col_count": 2,
                        "table_type": "general",
                    }
                ],
            },
        ],
        "numeric_data": [
            {"page": 1, "table": 1, "header": "Q3 2025", "value": "$12,450,000", "numeric_value": 12450000.0, "data_type": "currency"},
            {"page": 1, "table": 1, "header": "Growth", "value": "15.3%", "numeric_value": 15.3, "data_type": "percent"},
            {"page": 1, "table": 1, "header": "Q3 2025", "value": "$8,750,000", "numeric_value": 8750000.0, "data_type": "currency"},
            {"page": 1, "table": 1, "header": "Growth", "value": "(2.1%)", "numeric_value": -2.1, "data_type": "percent"},
            {"page": 2, "table": 2, "header": "Units Sold", "value": "1,250", "numeric_value": 1250.0, "data_type": "number"},
        ],
        "total_tables": 2,
        "total_numeric_values": 5,
    }


@pytest.fixture
def empty_extraction():
    """Extraction result with no tables."""
    return {
        "metadata": {"title": "", "author": "", "page_count": "1"},
        "pages": [],
        "numeric_data": [],
        "total_tables": 0,
        "total_numeric_values": 0,
    }


# ---------------------------------------------------------------------------
# Per-table clean CSV
# ---------------------------------------------------------------------------


class TestGenerateTableCsv:
    def test_numeric_columns_use_parsed_value(self, sample_extraction):
        table = sample_extraction["pages"][0]["tables"][0]
        csv_out = generate_table_csv(table, 1, "test.pdf")
        reader = csv.DictReader(io.StringIO(csv_out))
        rows = list(reader)
        assert rows[0]["Q3 2025"] == "12450000.0"

    def test_text_columns_use_raw_value(self, sample_extraction):
        table = sample_extraction["pages"][0]["tables"][0]
        csv_out = generate_table_csv(table, 1, "test.pdf")
        reader = csv.DictReader(io.StringIO(csv_out))
        rows = list(reader)
        assert rows[0]["Segment"] == "North America"

    def test_clean_columns_only(self, sample_extraction):
        """Per-table CSV should only have the original column headers."""
        table = sample_extraction["pages"][0]["tables"][0]
        csv_out = generate_table_csv(table, 1, "test.pdf")
        reader = csv.DictReader(io.StringIO(csv_out))
        fieldnames = reader.fieldnames
        assert fieldnames == ["Segment", "Q3 2025", "Growth"]
        assert "dtype" not in " ".join(fieldnames)
        assert "(raw)" not in " ".join(fieldnames)

    def test_no_section_markers(self, sample_extraction):
        table = sample_extraction["pages"][0]["tables"][0]
        csv_out = generate_table_csv(table, 1, "test.pdf")
        assert "---" not in csv_out

    def test_no_blank_rows(self, sample_extraction):
        table = sample_extraction["pages"][0]["tables"][0]
        csv_out = generate_table_csv(table, 1, "test.pdf")
        lines = [line for line in csv_out.strip().split("\n") if line.strip()]
        # Header + 2 data rows = 3 lines
        assert len(lines) == 3

    def test_negative_values_preserved(self, sample_extraction):
        table = sample_extraction["pages"][0]["tables"][0]
        csv_out = generate_table_csv(table, 1, "test.pdf")
        reader = csv.DictReader(io.StringIO(csv_out))
        rows = list(reader)
        assert float(rows[1]["Growth"]) == -2.1

    def test_parseable_by_csv_reader(self, sample_extraction):
        table = sample_extraction["pages"][0]["tables"][0]
        csv_out = generate_table_csv(table, 1, "test.pdf")
        reader = csv.reader(io.StringIO(csv_out))
        all_rows = list(reader)
        assert len(all_rows) == 3  # header + 2 data
        assert len(all_rows[0]) == 3  # original 3 columns, no companions
        # All rows same column count
        assert all(len(r) == len(all_rows[0]) for r in all_rows)


# ---------------------------------------------------------------------------
# Combined tidy CSV
# ---------------------------------------------------------------------------


class TestGenerateCombinedCsv:
    def test_all_rows_present(self, sample_extraction):
        csv_out = generate_combined_csv(sample_extraction, "test.pdf")
        reader = csv.DictReader(io.StringIO(csv_out))
        rows = list(reader)
        # Table 1 has 2 rows x 3 cols = 6 cell observations
        # Table 2 has 1 row x 2 cols = 2 cell observations
        assert len(rows) == 8

    def test_metadata_columns_present(self, sample_extraction):
        csv_out = generate_combined_csv(sample_extraction, "test.pdf")
        reader = csv.DictReader(io.StringIO(csv_out))
        rows = list(reader)
        assert rows[0]["source_file"] == "test.pdf"
        assert rows[0]["page"] == "1"
        assert rows[0]["table_id"] == "table_001"
        assert rows[0]["table_type"] == "financial"

    def test_column_headers(self, sample_extraction):
        csv_out = generate_combined_csv(sample_extraction, "test.pdf")
        reader = csv.DictReader(io.StringIO(csv_out))
        fieldnames = reader.fieldnames
        expected = [
            "source_file", "page", "table_id", "table_type",
            "row_number", "column_name", "raw_value", "parsed_value", "data_type",
        ]
        assert fieldnames == expected

    def test_parsed_values(self, sample_extraction):
        csv_out = generate_combined_csv(sample_extraction, "test.pdf")
        reader = csv.DictReader(io.StringIO(csv_out))
        rows = list(reader)
        # Find a currency cell
        currency_rows = [r for r in rows if r["data_type"] == "currency"]
        assert len(currency_rows) > 0
        assert currency_rows[0]["parsed_value"] == "12450000.0"
        assert currency_rows[0]["raw_value"] == "$12,450,000"

    def test_text_cells_have_empty_parsed(self, sample_extraction):
        csv_out = generate_combined_csv(sample_extraction, "test.pdf")
        reader = csv.DictReader(io.StringIO(csv_out))
        rows = list(reader)
        text_rows = [r for r in rows if r["data_type"] == "text"]
        for r in text_rows:
            assert r["parsed_value"] == ""

    def test_empty_extraction(self, empty_extraction):
        csv_out = generate_combined_csv(empty_extraction, "empty.pdf")
        reader = csv.DictReader(io.StringIO(csv_out))
        rows = list(reader)
        assert len(rows) == 0

    def test_no_section_markers(self, sample_extraction):
        csv_out = generate_combined_csv(sample_extraction, "test.pdf")
        assert "---" not in csv_out


# ---------------------------------------------------------------------------
# Structured JSON
# ---------------------------------------------------------------------------


class TestGenerateJsonExport:
    def test_valid_json(self, sample_extraction):
        json_out = generate_json_export(sample_extraction, "test.pdf")
        result = json.loads(json_out)
        assert isinstance(result, dict)

    def test_structure(self, sample_extraction):
        result = json.loads(generate_json_export(sample_extraction, "test.pdf"))
        assert result["source_file"] == "test.pdf"
        assert "extracted_at" in result
        assert "metadata" in result
        assert "summary" in result
        assert len(result["tables"]) == 2

    def test_data_uses_numeric_values(self, sample_extraction):
        result = json.loads(generate_json_export(sample_extraction, "test.pdf"))
        t1_data = result["tables"][0]["data"]
        assert t1_data[0]["Q3 2025"] == 12450000.0
        assert isinstance(t1_data[0]["Q3 2025"], float)

    def test_data_raw_preserves_formatting(self, sample_extraction):
        result = json.loads(generate_json_export(sample_extraction, "test.pdf"))
        t1_raw = result["tables"][0]["data_raw"]
        assert t1_raw[0]["Q3 2025"] == "$12,450,000"

    def test_column_schema(self, sample_extraction):
        result = json.loads(generate_json_export(sample_extraction, "test.pdf"))
        columns = result["tables"][0]["columns"]
        assert len(columns) == 3
        seg_col = [c for c in columns if c["name"] == "Segment"][0]
        assert seg_col["dominant_type"] == "text"
        q3_col = [c for c in columns if c["name"] == "Q3 2025"][0]
        assert q3_col["dominant_type"] == "currency"

    def test_negative_numeric_values(self, sample_extraction):
        result = json.loads(generate_json_export(sample_extraction, "test.pdf"))
        t1_data = result["tables"][0]["data"]
        assert t1_data[1]["Growth"] == -2.1

    def test_empty_extraction(self, empty_extraction):
        result = json.loads(generate_json_export(empty_extraction, "empty.pdf"))
        assert result["tables"] == []
        assert result["summary"]["total_tables"] == 0


# ---------------------------------------------------------------------------
# Excel workbook
# ---------------------------------------------------------------------------


class TestGenerateExcel:
    def test_excel_is_valid(self, sample_extraction):
        import openpyxl

        excel_bytes = generate_excel(sample_extraction, "test.pdf")
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        assert "Summary" in wb.sheetnames

    def test_excel_sheet_count(self, sample_extraction):
        import openpyxl

        excel_bytes = generate_excel(sample_extraction, "test.pdf")
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        # Summary + Data Quality + 2 table sheets = 4
        assert len(wb.sheetnames) == 4
        assert "Summary" in wb.sheetnames
        assert "Data Quality" in wb.sheetnames

    def test_excel_numeric_values(self, sample_extraction):
        import openpyxl

        excel_bytes = generate_excel(sample_extraction, "test.pdf")
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        # First table sheet (after Summary and Data Quality)
        ws = wb[wb.sheetnames[2]]
        # Row 2 (first data row), Col 2 (Q3 2025) should be numeric
        val = ws.cell(row=2, column=2).value
        assert isinstance(val, (int, float))
        assert val == 12450000.0

    def test_excel_percent_divided_by_100(self, sample_extraction):
        import openpyxl

        excel_bytes = generate_excel(sample_extraction, "test.pdf")
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        # First table sheet (after Summary and Data Quality)
        ws = wb[wb.sheetnames[2]]
        # Row 2, Col 3 (Growth = 15.3%) should be stored as 0.153
        val = ws.cell(row=2, column=3).value
        assert abs(val - 0.153) < 0.0001

    def test_excel_summary_sheet(self, sample_extraction):
        import openpyxl

        excel_bytes = generate_excel(sample_extraction, "test.pdf")
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb["Summary"]
        # Should contain source filename
        values = [ws.cell(row=r, column=2).value for r in range(1, 20)]
        assert "test.pdf" in values


# ---------------------------------------------------------------------------
# ZIP bundle
# ---------------------------------------------------------------------------


class TestGenerateZipBundle:
    def test_zip_is_valid(self, sample_extraction):
        zip_bytes = generate_zip_bundle(
            sample_extraction, "report.pdf", "# Markdown", "a,b\n1,2",
        )
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
        assert len(zf.namelist()) > 0

    def test_zip_contains_expected_files(self, sample_extraction):
        zip_bytes = generate_zip_bundle(
            sample_extraction, "report.pdf", "# Markdown", "a,b\n1,2",
        )
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
        names = zf.namelist()
        assert any("combined_tidy.csv" in n for n in names)
        assert any("data.json" in n for n in names)
        assert any("table_001" in n for n in names)
        assert any("table_002" in n for n in names)
        assert any(".md" in n for n in names)
        assert any(".xlsx" in n for n in names)

    def test_zip_csv_is_clean(self, sample_extraction):
        zip_bytes = generate_zip_bundle(
            sample_extraction, "report.pdf", "# Markdown", "a,b\n1,2",
        )
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
        # Find a per-table CSV
        csv_name = [n for n in zf.namelist() if "table_001" in n][0]
        csv_content = zf.read(csv_name).decode("utf-8")
        assert "---" not in csv_content
        reader = csv.reader(io.StringIO(csv_content))
        rows = list(reader)
        assert len(rows) == 3  # header + 2 data rows

    def test_zip_json_is_valid(self, sample_extraction):
        zip_bytes = generate_zip_bundle(
            sample_extraction, "report.pdf", "# Markdown", "a,b\n1,2",
        )
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
        json_name = [n for n in zf.namelist() if "data.json" in n][0]
        data = json.loads(zf.read(json_name))
        assert len(data["tables"]) == 2

    def test_zip_contains_quality_report(self, sample_extraction):
        zip_bytes = generate_zip_bundle(
            sample_extraction, "report.pdf", "# Markdown", "a,b\n1,2",
        )
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
        names = zf.namelist()
        assert any("data_quality.json" in n for n in names)
        quality_name = [n for n in names if "data_quality.json" in n][0]
        quality = json.loads(zf.read(quality_name))
        assert "tables" in quality
        assert "overall_quality" in quality


# ---------------------------------------------------------------------------
# Data quality report
# ---------------------------------------------------------------------------


class TestDataQualityReport:
    def test_basic_structure(self, sample_extraction):
        report = generate_data_quality_report(sample_extraction, "test.pdf")
        assert report["source_file"] == "test.pdf"
        assert report["total_tables"] == 2
        assert "overall_quality" in report
        assert len(report["tables"]) == 2

    def test_column_stats(self, sample_extraction):
        report = generate_data_quality_report(sample_extraction, "test.pdf")
        table_1 = report["tables"][0]
        assert len(table_1["column_stats"]) == 3
        # Segment column is all text
        seg = table_1["column_stats"][0]
        assert seg["column_name"] == "Segment"
        assert seg["dominant_type"] == "text"
        assert seg["missing"] == 0

    def test_numeric_stats(self, sample_extraction):
        report = generate_data_quality_report(sample_extraction, "test.pdf")
        table_1 = report["tables"][0]
        q3_col = [c for c in table_1["column_stats"] if c["column_name"] == "Q3 2025"][0]
        assert "min" in q3_col
        assert "max" in q3_col
        assert "mean" in q3_col
        assert q3_col["min"] == 8750000.0
        assert q3_col["max"] == 12450000.0

    def test_empty_extraction(self, empty_extraction):
        report = generate_data_quality_report(empty_extraction, "empty.pdf")
        assert report["total_tables"] == 0
        assert report["overall_quality"] == "good"

    def test_json_includes_quality(self, sample_extraction):
        json_out = generate_json_export(sample_extraction, "test.pdf")
        data = json.loads(json_out)
        assert "data_quality" in data
        assert data["data_quality"]["total_tables"] == 2
