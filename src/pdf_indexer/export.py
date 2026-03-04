"""Analyst-friendly export formats for extracted PDF table data.

Generates clean CSV, JSON, Excel, and ZIP outputs that can be loaded
directly in Jupyter Notebooks, R, and Python without manual cleanup.
"""

from __future__ import annotations

import csv
import io
import json
import zipfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path


def _dominant_column_type(col_idx: int, rows: list[list[dict]]) -> str:
    """Return the most common data_type for a column, or 'text'."""
    types: list[str] = []
    for row in rows:
        if col_idx < len(row) and row[col_idx]["value"].strip():
            types.append(row[col_idx]["data_type"])
    if not types:
        return "text"
    counter = Counter(types)
    most_common, _ = counter.most_common(1)[0]
    return most_common


def _is_numeric_column(col_idx: int, rows: list[list[dict]]) -> bool:
    """Return True if all non-empty cells in this column have a numeric_value."""
    has_value = False
    for row in rows:
        if col_idx < len(row):
            cell = row[col_idx]
            if cell["value"].strip():
                has_value = True
                if cell["numeric_value"] is None:
                    return False
    return has_value


def _empty_cell() -> dict:
    return {"value": "", "data_type": "text", "numeric_value": None}


def generate_table_csv(table_info: dict, page_num: int, source_filename: str) -> str:
    """Generate a clean CSV for a single table.

    Uses only the original column headers. Numeric columns contain parsed
    values (floats), text columns contain the original string. No companion
    columns — the schema is kept in the JSON export.

    Args:
        table_info: Table dict from extraction result.
        page_num: Page number the table was found on.
        source_filename: Original PDF filename.

    Returns:
        CSV string with headers + data rows only. No section markers.
    """
    headers = table_info["headers"]
    rows = table_info["rows"]
    col_numeric = [_is_numeric_column(i, rows) for i in range(len(headers))]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)

    for row in rows:
        out_row: list[str] = []
        for i in range(len(headers)):
            cell = row[i] if i < len(row) else _empty_cell()
            if col_numeric[i] and cell["numeric_value"] is not None:
                out_row.append(str(cell["numeric_value"]))
            else:
                out_row.append(cell["value"])
        writer.writerow(out_row)

    return output.getvalue()


def generate_combined_csv(extraction: dict, source_filename: str) -> str:
    """Generate a tidy long-format CSV combining all tables.

    Each row represents one cell observation with metadata columns:
    source_file, page, table_id, table_type, row_number, column_name,
    raw_value, parsed_value, data_type

    This format is universally tidy and can be pivoted as needed.

    Args:
        extraction: Full extraction result dict.
        source_filename: Original PDF filename.

    Returns:
        CSV string in long/melted format.
    """
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "source_file", "page", "table_id", "table_type",
        "row_number", "column_name", "raw_value", "parsed_value", "data_type",
    ])

    for page_info in extraction["pages"]:
        page_num = page_info["page_number"]
        for table_info in page_info["tables"]:
            table_id = f"table_{table_info['table_number']:03d}"
            table_type = table_info["table_type"]
            headers = table_info["headers"]

            for row_idx, row in enumerate(table_info["rows"], start=1):
                for col_idx, header in enumerate(headers):
                    cell = row[col_idx] if col_idx < len(row) else _empty_cell()
                    parsed = (
                        str(cell["numeric_value"])
                        if cell["numeric_value"] is not None
                        else ""
                    )
                    writer.writerow([
                        source_filename,
                        page_num,
                        table_id,
                        table_type,
                        row_idx,
                        header,
                        cell["value"],
                        parsed,
                        cell["data_type"],
                    ])

    return output.getvalue()


def generate_data_quality_report(extraction: dict, source_filename: str) -> dict:
    """Generate a data quality summary for extracted tables.

    Returns a dict with per-table quality metrics that a data analyst
    would check before beginning analysis.
    """
    tables_quality: list[dict] = []

    for page_info in extraction["pages"]:
        page_num = page_info["page_number"]
        for table_info in page_info["tables"]:
            headers = table_info["headers"]
            rows = table_info["rows"]
            n_rows = len(rows)
            n_cols = len(headers)

            # Per-column analysis
            col_stats: list[dict] = []
            for col_idx, header in enumerate(headers):
                values = []
                non_empty = 0
                types_seen: Counter = Counter()
                numeric_vals: list[float] = []

                for row in rows:
                    cell = row[col_idx] if col_idx < len(row) else _empty_cell()
                    val = cell["value"].strip()
                    values.append(val)
                    if val:
                        non_empty += 1
                    types_seen[cell["data_type"]] += 1
                    if cell["numeric_value"] is not None:
                        numeric_vals.append(cell["numeric_value"])

                missing_count = n_rows - non_empty
                missing_pct = round(missing_count / n_rows * 100, 1) if n_rows > 0 else 0.0
                is_mixed_type = len([t for t, c in types_seen.items()
                                     if c > 0 and t != "text"]) > 1

                col_stat: dict = {
                    "column_name": header,
                    "non_empty": non_empty,
                    "missing": missing_count,
                    "missing_pct": missing_pct,
                    "dominant_type": _dominant_column_type(col_idx, rows),
                    "types_found": dict(types_seen),
                    "mixed_types": is_mixed_type,
                }

                if numeric_vals:
                    col_stat["min"] = min(numeric_vals)
                    col_stat["max"] = max(numeric_vals)
                    col_stat["mean"] = round(sum(numeric_vals) / len(numeric_vals), 2)

                # Detect duplicates
                unique_count = len(set(values))
                col_stat["unique_values"] = unique_count
                col_stat["all_unique"] = unique_count == n_rows

                col_stats.append(col_stat)

            # Table-level issues
            issues: list[str] = []
            for cs in col_stats:
                if cs["missing_pct"] > 20:
                    issues.append(
                        f"Column '{cs['column_name']}' has {cs['missing_pct']}% missing values"
                    )
                if cs["mixed_types"]:
                    issues.append(
                        f"Column '{cs['column_name']}' has mixed data types: "
                        f"{cs['types_found']}"
                    )

            # Check for empty/duplicate headers
            if any(not h.strip() for h in headers):
                issues.append("Table has empty column headers")
            if len(set(headers)) < len(headers):
                issues.append("Table has duplicate column headers")

            tables_quality.append({
                "table_id": table_info["table_number"],
                "page": page_num,
                "table_type": table_info["table_type"],
                "rows": n_rows,
                "columns": n_cols,
                "column_stats": col_stats,
                "issues": issues,
                "quality_score": "good" if not issues else "needs_review",
            })

    total_issues = sum(len(t["issues"]) for t in tables_quality)
    return {
        "source_file": source_filename,
        "total_tables": len(tables_quality),
        "tables_with_issues": sum(1 for t in tables_quality if t["issues"]),
        "total_issues": total_issues,
        "overall_quality": "good" if total_issues == 0 else "needs_review",
        "tables": tables_quality,
    }


def generate_json_export(extraction: dict, source_filename: str) -> str:
    """Generate structured JSON for programmatic loading.

    Includes both parsed numeric data (for computation) and raw formatted
    data (for display/validation), plus a column schema acting as a
    data dictionary, and a data quality report.

    Args:
        extraction: Full extraction result dict.
        source_filename: Original PDF filename.

    Returns:
        JSON string.
    """
    result: dict = {
        "source_file": source_filename,
        "extracted_at": datetime.now(timezone.utc).isoformat(),
        "metadata": extraction["metadata"],
        "summary": {
            "total_tables": extraction["total_tables"],
            "total_numeric_values": extraction["total_numeric_values"],
        },
        "tables": [],
    }

    for page_info in extraction["pages"]:
        page_num = page_info["page_number"]
        for table_info in page_info["tables"]:
            headers = table_info["headers"]
            rows = table_info["rows"]

            # Build column schema
            columns = []
            for i, h in enumerate(headers):
                col_type = _dominant_column_type(i, rows)
                sample_values = []
                for row in rows[:3]:
                    if i < len(row):
                        sample_values.append(row[i]["value"])
                columns.append({
                    "name": h,
                    "dominant_type": col_type,
                    "sample_values": sample_values,
                })

            # Build data arrays
            data_parsed: list[dict] = []
            data_raw: list[dict] = []

            for row in rows:
                parsed_row: dict = {}
                raw_row: dict = {}
                for i, h in enumerate(headers):
                    cell = row[i] if i < len(row) else _empty_cell()
                    raw_row[h] = cell["value"]
                    if cell["numeric_value"] is not None:
                        parsed_row[h] = cell["numeric_value"]
                    else:
                        parsed_row[h] = cell["value"]
                data_parsed.append(parsed_row)
                data_raw.append(raw_row)

            table_entry: dict = {
                "table_id": table_info["table_number"],
                "page": page_num,
                "table_type": table_info["table_type"],
                "row_count": table_info["row_count"],
                "col_count": table_info["col_count"],
                "columns": columns,
                "data": data_parsed,
                "data_raw": data_raw,
            }

            # Include context if available
            if table_info.get("context_before"):
                table_entry["context_before"] = table_info["context_before"]
            if table_info.get("context_after"):
                table_entry["context_after"] = table_info["context_after"]

            result["tables"].append(table_entry)

    # Append data quality report
    result["data_quality"] = generate_data_quality_report(extraction, source_filename)

    return json.dumps(result, indent=2, ensure_ascii=False)


def generate_excel(extraction: dict, source_filename: str) -> bytes:
    """Generate an Excel workbook with one sheet per table plus a summary.

    Args:
        extraction: Full extraction result dict.
        source_filename: Original PDF filename.

    Returns:
        Bytes of the .xlsx file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()

    # --- Summary sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["PDF Table Extraction Summary"])
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.append([])
    ws_summary.append(["Source File", source_filename])
    ws_summary.append(["Extracted At", datetime.now(timezone.utc).isoformat()])

    meta = extraction["metadata"]
    ws_summary.append(["Title", meta.get("title", "")])
    ws_summary.append(["Author", meta.get("author", "")])
    ws_summary.append(["Pages", meta.get("page_count", "")])
    ws_summary.append(["Total Tables", extraction["total_tables"]])
    ws_summary.append(["Total Numeric Values", extraction["total_numeric_values"]])
    ws_summary.append([])

    # Table index
    ws_summary.append(["Table Index"])
    ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True, size=12)
    ws_summary.append(["Table ID", "Page", "Type", "Rows", "Columns", "Context"])
    for cell in ws_summary[ws_summary.max_row]:
        cell.font = Font(bold=True)

    for page_info in extraction["pages"]:
        for table_info in page_info["tables"]:
            context = table_info.get("context_before", "")
            if context and len(context) > 80:
                context = context[:80] + "..."
            ws_summary.append([
                table_info["table_number"],
                page_info["page_number"],
                table_info["table_type"],
                table_info["row_count"],
                table_info["col_count"],
                context,
            ])

    # Auto-fit summary columns
    for col in ws_summary.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws_summary.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # --- Data Quality sheet ---
    quality = generate_data_quality_report(extraction, source_filename)
    ws_quality = wb.create_sheet(title="Data Quality")
    ws_quality.append(["Data Quality Report"])
    ws_quality["A1"].font = Font(bold=True, size=14)
    ws_quality.append([])
    ws_quality.append([
        "Overall Quality", quality["overall_quality"].upper(),
        "", f"{quality['total_issues']} issue(s) across "
        f"{quality['tables_with_issues']} table(s)",
    ])
    if quality["overall_quality"] == "good":
        ws_quality.cell(row=3, column=2).font = Font(bold=True, color="228B22")
    else:
        ws_quality.cell(row=3, column=2).font = Font(bold=True, color="FF4500")
    ws_quality.append([])

    red_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")

    for tq in quality["tables"]:
        ws_quality.append([f"Table {tq['table_id']} (Page {tq['page']}, {tq['table_type']})"])
        ws_quality.cell(row=ws_quality.max_row, column=1).font = Font(bold=True, size=11)
        ws_quality.append([
            "Column", "Type", "Non-Empty", "Missing", "Missing %",
            "Unique", "Min", "Max", "Mean",
        ])
        for cell in ws_quality[ws_quality.max_row]:
            cell.font = Font(bold=True)

        for cs in tq["column_stats"]:
            row_data = [
                cs["column_name"],
                cs["dominant_type"],
                cs["non_empty"],
                cs["missing"],
                cs["missing_pct"],
                cs["unique_values"],
                cs.get("min", ""),
                cs.get("max", ""),
                cs.get("mean", ""),
            ]
            ws_quality.append(row_data)
            if cs["missing_pct"] > 20 or cs["mixed_types"]:
                for cell in ws_quality[ws_quality.max_row]:
                    cell.fill = red_fill

        if tq["issues"]:
            ws_quality.append(["Issues:"])
            ws_quality.cell(row=ws_quality.max_row, column=1).font = Font(
                bold=True, color="FF4500",
            )
            for issue in tq["issues"]:
                ws_quality.append(["", issue])
        ws_quality.append([])

    for col in ws_quality.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws_quality.column_dimensions[col_letter].width = min(max_len + 2, 40)

    # --- One sheet per table ---
    for page_info in extraction["pages"]:
        page_num = page_info["page_number"]
        for table_info in page_info["tables"]:
            tnum = table_info["table_number"]
            sheet_name = f"Table {tnum} (P{page_num})"
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            ws = wb.create_sheet(title=sheet_name)
            headers = table_info["headers"]
            rows = table_info["rows"]

            # Determine column types
            col_numeric = [_is_numeric_column(i, rows) for i in range(len(headers))]
            col_types = [_dominant_column_type(i, rows) for i in range(len(headers))]

            # Write headers
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Write data rows
            for row in rows:
                out_row = []
                for i in range(len(headers)):
                    cell = row[i] if i < len(row) else _empty_cell()
                    if col_numeric[i] and cell["numeric_value"] is not None:
                        if cell["data_type"] == "percent":
                            out_row.append(cell["numeric_value"] / 100.0)
                        else:
                            out_row.append(cell["numeric_value"])
                    else:
                        out_row.append(cell["value"])
                ws.append(out_row)

            # Apply number formats to data cells
            for col_idx in range(len(headers)):
                if not col_numeric[col_idx]:
                    continue
                fmt = None
                if col_types[col_idx] == "currency":
                    fmt = '$#,##0.00'
                elif col_types[col_idx] == "percent":
                    fmt = '0.0%'
                elif col_types[col_idx] == "number":
                    fmt = '#,##0.##'

                if fmt:
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx + 1)
                        cell.number_format = fmt

            # Auto-fit column widths
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

            # Add auto-filter on header row
            ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_zip_bundle(
    extraction: dict,
    source_filename: str,
    markdown: str,
    legacy_csv: str,
) -> bytes:
    """Generate a ZIP bundle containing all export formats.

    Contents:
        tables/table_NNN_pN.csv  - clean per-table CSVs
        combined_tidy.csv        - long-format combined dataset
        data.json                - structured JSON with quality report
        report.xlsx              - Excel workbook with quality sheet
        report.md                - markdown report
        data_quality.json        - standalone quality report

    Args:
        extraction: Full extraction result dict.
        source_filename: Original PDF filename.
        markdown: Existing markdown report string.
        legacy_csv: Legacy CSV string for backward compatibility.

    Returns:
        Bytes of the .zip file.
    """
    buf = io.BytesIO()
    stem = Path(source_filename).stem
    prefix = f"{stem}_export"

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Per-table clean CSVs
        for page_info in extraction["pages"]:
            page_num = page_info["page_number"]
            for table_info in page_info["tables"]:
                tnum = table_info["table_number"]
                csv_name = f"table_{tnum:03d}_p{page_num}.csv"
                csv_content = generate_table_csv(table_info, page_num, source_filename)
                zf.writestr(f"{prefix}/tables/{csv_name}", csv_content)

        # Combined tidy CSV
        tidy = generate_combined_csv(extraction, source_filename)
        zf.writestr(f"{prefix}/combined_tidy.csv", tidy)

        # Structured JSON (includes quality report)
        json_content = generate_json_export(extraction, source_filename)
        zf.writestr(f"{prefix}/data.json", json_content)

        # Standalone data quality report
        quality = generate_data_quality_report(extraction, source_filename)
        zf.writestr(
            f"{prefix}/data_quality.json",
            json.dumps(quality, indent=2, ensure_ascii=False),
        )

        # Excel workbook (includes quality sheet)
        try:
            excel_bytes = generate_excel(extraction, source_filename)
            zf.writestr(f"{prefix}/{stem}_tables.xlsx", excel_bytes)
        except ImportError:
            pass

        # Markdown report
        zf.writestr(f"{prefix}/{stem}_tables.md", markdown)

    return buf.getvalue()
